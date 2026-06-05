#!/usr/bin/env python3
"""
init_excel_workbook.py
======================
Initialize the consolidated.xlsx workbook with the proper sheet structure.

This script creates an empty Excel file with the following sheets:
  - Raw Data: Where you paste new/incoming data
  - Historical Data: Auto-populated with all historical records (deduplicated)
  - Insights: Auto-generated metrics and KPIs
  - Metadata: Tracking information and processing metadata

Run this once to set up the workbook structure, then use process_data.py for updates.

Usage:
    python init_excel_workbook.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


DATA_FOLDER = "Data"
CONSOLIDATED_FILE = os.path.join(DATA_FOLDER, "consolidated.xlsx")


def create_header_style():
    """Create a header cell style."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }


def style_cell(cell, style):
    """Apply style to a cell."""
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]


def create_raw_data_sheet(ws):
    """Create Raw Data sheet with expected columns."""
    ws.title = "Raw Data"
    
    headers = [
        "Case No.",
        "Creation Date",
        "Closing Date",
        "Workstream",
        "Priority",
        "Error Category",
        "Days_AtCustomer",
        "Days_AtSAP",
        "Flag_HowToOrConsulting",
        "Flag_PriorityInflation",
        "Flag_CustomerDelay",
        "Flag_ComponentMismatch",
        "Flag_Reopened",
        "Flag_AutoConfirmed",
        "Flag_TotalCoaching",
    ]
    
    header_style = create_header_style()
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_cell(cell, header_style)
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Add instruction row
    instruction_cell = ws.cell(row=2, column=1, value="← Paste your data below this row")
    instruction_cell.font = Font(italic=True, color="999999", size=9)
    
    print(f"✓ Created 'Raw Data' sheet with {len(headers)} columns")


def create_historical_data_sheet(ws):
    """Create Historical Data sheet."""
    ws.title = "Historical Data"
    
    headers = [
        "Case No.",
        "Creation Date",
        "Closing Date",
        "Workstream",
        "Priority",
        "Error Category",
        "Days_AtCustomer",
        "Days_AtSAP",
        "Flag_HowToOrConsulting",
        "Flag_PriorityInflation",
        "Flag_CustomerDelay",
        "Flag_ComponentMismatch",
        "Flag_Reopened",
        "Flag_AutoConfirmed",
        "Flag_TotalCoaching",
        "_processed_date",
    ]
    
    header_style = create_header_style()
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_cell(cell, header_style)
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Add note
    note_cell = ws.cell(row=2, column=1, value="← Historical data (auto-populated and deduplicated)")
    note_cell.font = Font(italic=True, color="999999", size=9)
    
    print(f"✓ Created 'Historical Data' sheet with {len(headers)} columns")


def create_insights_sheet(ws):
    """Create Insights sheet with metrics headers."""
    ws.title = "Insights"
    
    insights_headers = [
        ["OVERALL METRICS", ""],
        ["Total Cases", "0"],
        ["Date Range", "N/A"],
        ["Flagged Cases", "0"],
        ["", ""],
        ["COACHING FLAGS", "Count", "Percentage"],
        ["How-To / Consulting", "0", "0%"],
        ["Priority Inflation", "0", "0%"],
        ["Customer Delay", "0", "0%"],
        ["Component Mismatch", "0", "0%"],
        ["Reopened", "0", "0%"],
        ["Auto-Confirmed", "0", "0%"],
        ["", ""],
        ["TIME METRICS", "Days"],
        ["Avg Days @ Customer", "0"],
        ["Avg Days @ SAP", "0"],
        ["", ""],
        ["TOP COACHING ISSUE", ""],
        ["N/A", "0"],
    ]
    
    header_style = create_header_style()
    
    for row_idx, row_data in enumerate(insights_headers, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style section headers (all caps)
            if isinstance(value, str) and value.isupper() and len(value) > 3 and col_idx == 1:
                style_cell(cell, header_style)
            
            ws.column_dimensions[chr(64 + col_idx)].width = 25
    
    print(f"✓ Created 'Insights' sheet with KPI template")


def create_metadata_sheet(ws):
    """Create Metadata sheet for tracking."""
    ws.title = "Metadata"
    
    metadata = [
        ["Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Rows Processed", "0"],
        ["Data Quality", "Ready"],
        ["Processing Status", "Initialized"],
        ["", ""],
        ["INSTRUCTIONS", ""],
        ["1. Add raw data", "Paste data to 'Raw Data' sheet"],
        ["2. Run processing", "Execute: python process_data.py"],
        ["3. Review insights", "Check 'Insights' sheet for KPIs"],
        ["4. Check history", "View 'Historical Data' for all records"],
        ["5. Generate dashboard", "Execute: python build_dashboard.py"],
    ]
    
    header_style = create_header_style()
    
    for row_idx, row_data in enumerate(metadata, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style section header
            if value == "INSTRUCTIONS":
                style_cell(cell, header_style)
            
            ws.column_dimensions[chr(64 + col_idx)].width = 30
    
    print(f"✓ Created 'Metadata' sheet with tracking info")


def main():
    """Main initialization."""
    print("=" * 60)
    print("Excel Workbook Initialization")
    print("=" * 60)
    
    # Create data folder if it doesn't exist
    os.makedirs(DATA_FOLDER, exist_ok=True)
    
    # Check if file already exists
    if os.path.exists(CONSOLIDATED_FILE):
        response = input(f"\n⚠ {CONSOLIDATED_FILE} already exists. Overwrite? (y/n): ")
        if response.lower() != 'y':
            print("Cancelled.")
            return
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    print(f"\nCreating {CONSOLIDATED_FILE}...\n")
    
    # Create sheets
    create_raw_data_sheet(wb.create_sheet())
    create_historical_data_sheet(wb.create_sheet())
    create_insights_sheet(wb.create_sheet())
    create_metadata_sheet(wb.create_sheet())
    
    # Save
    wb.save(CONSOLIDATED_FILE)
    
    print(f"\n{'=' * 60}")
    print(f"✓ Created: {CONSOLIDATED_FILE}")
    print(f"{'=' * 60}")
    print("\nNext steps:")
    print("  1. Open the Excel file and add your data to 'Raw Data' sheet")
    print("  2. Run: python process_data.py")
    print("  3. Run: python build_dashboard.py")
    print("  4. Open: output/index.html")
    print()


if __name__ == "__main__":
    main()
