#!/usr/bin/env python3
"""
process_data.py
===============
Automated data processing script for OSS Dashboard.

This script:
1. Reads raw data from the 'Raw Data' sheet in consolidated.xlsx
2. Appends it to the 'Historical Data' sheet with timestamps
3. Regenerates the 'Insights' sheet with current metrics
4. Maintains data history for trend analysis

Usage:
    python process_data.py

Can be triggered via GitHub Actions on a schedule or manually.
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# Configuration
DATA_FOLDER = "Data"
CONSOLIDATED_FILE = os.path.join(DATA_FOLDER, "consolidated.xlsx")
BACKUP_FOLDER = os.path.join(DATA_FOLDER, "backups")
LOG_FILE = os.path.join(DATA_FOLDER, "process.log")

# Sheet names
SHEET_RAW_DATA = "Raw Data"
SHEET_HISTORICAL = "Historical Data"
SHEET_INSIGHTS = "Insights"
SHEET_METADATA = "Metadata"

# Flag columns (matching build_dashboard.py)
FLAG_COLUMNS = [
    "Flag_HowToOrConsulting",
    "Flag_PriorityInflation",
    "Flag_CustomerDelay",
    "Flag_ComponentMismatch",
    "Flag_Reopened",
    "Flag_AutoConfirmed",
    "Flag_TotalCoaching",
]

FLAG_LABELS = {
    "Flag_HowToOrConsulting": "How-To / Consulting",
    "Flag_PriorityInflation": "Priority Inflation",
    "Flag_CustomerDelay": "Customer Delay",
    "Flag_ComponentMismatch": "Component Mismatch",
    "Flag_Reopened": "Reopened",
    "Flag_AutoConfirmed": "Auto-Confirmed",
    "Flag_TotalCoaching": "Total Coaching",
}


def log(msg):
    """Log message to console and log file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    print(log_msg)
    os.makedirs(DATA_FOLDER, exist_ok=True)
    with open(LOG_FILE, "a") as f:
        f.write(log_msg + "\n")


def ensure_sheets_exist(file_path):
    """Ensure all required sheets exist in the Excel file."""
    os.makedirs(DATA_FOLDER, exist_ok=True)
    
    if not os.path.exists(file_path):
        log(f"Creating new file: {file_path}")
        wb = load_workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create sheet structure
        for sheet_name in [SHEET_RAW_DATA, SHEET_HISTORICAL, SHEET_INSIGHTS, SHEET_METADATA]:
            wb.create_sheet(sheet_name)
        
        # Add metadata
        ws_meta = wb[SHEET_METADATA]
        ws_meta["A1"] = "Last Updated"
        ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws_meta["A2"] = "Total Rows Processed"
        ws_meta["B2"] = 0
        ws_meta["A3"] = "Data Quality"
        ws_meta["B3"] = "Ready"
        
        wb.save(file_path)
        log(f"Created file with empty sheets")
        return
    
    wb = load_workbook(file_path)
    
    # Ensure all sheets exist
    for sheet_name in [SHEET_RAW_DATA, SHEET_HISTORICAL, SHEET_INSIGHTS, SHEET_METADATA]:
        if sheet_name not in wb.sheetnames:
            log(f"Creating missing sheet: {sheet_name}")
            wb.create_sheet(sheet_name)
    
    wb.save(file_path)


def read_raw_data():
    """Read data from Raw Data sheet."""
    try:
        df = pd.read_excel(CONSOLIDATED_FILE, sheet_name=SHEET_RAW_DATA)
        if len(df) == 0:
            log("WARNING: Raw Data sheet is empty")
            return None
        log(f"Read {len(df)} rows from Raw Data sheet")
        return df
    except Exception as e:
        log(f"ERROR reading Raw Data: {e}")
        return None


def read_historical_data():
    """Read existing historical data."""
    try:
        df = pd.read_excel(CONSOLIDATED_FILE, sheet_name=SHEET_HISTORICAL)
        log(f"Read {len(df)} rows from Historical Data sheet")
        return df
    except Exception as e:
        log(f"WARNING reading Historical Data: {e} (might be first run)")
        return pd.DataFrame()


def clean_data(df):
    """Clean and standardize data types."""
    if df is None or len(df) == 0:
        return df
    
    # Clean flag columns
    for col in FLAG_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    
    # Clean numeric columns
    for col in ["Days_AtCustomer", "Days_AtSAP"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # Clean string columns
    for col in ["Workstream", "Priority", "Error Category"]:
        if col in df.columns:
            df[col] = df[col].fillna("Unknown").astype(str).str.strip()
    
    # Clean date columns
    for col in ["Creation Date", "Closing Date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    
    return df


def deduplicate_on_case_no(df):
    """Remove duplicates keeping the latest entry."""
    if df is None or len(df) == 0:
        return df
    
    if "Case No." not in df.columns:
        log("WARNING: 'Case No.' column not found - skipping deduplication")
        return df
    
    before = len(df)
    
    # Sort by Closing Date desc so latest is kept
    if "Closing Date" in df.columns:
        df["_sort_date"] = pd.to_datetime(df["Closing Date"], errors="coerce")
        df = df.sort_values("_sort_date", ascending=False, na_position="last")
        df = df.drop(columns=["_sort_date"])
    
    df = df.drop_duplicates(subset=["Case No."], keep="first")
    after = len(df)
    removed = before - after
    
    if removed > 0:
        log(f"Deduplicated: removed {removed} duplicate case(s)")
    
    return df.reset_index(drop=True)


def append_to_historical(raw_df, historical_df):
    """Append raw data to historical data with processing timestamp."""
    if raw_df is None or len(raw_df) == 0:
        log("No new data to append")
        return historical_df
    
    raw_df = raw_df.copy()
    raw_df["_processed_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if len(historical_df) == 0:
        combined = raw_df
        log(f"Initialized historical data with {len(raw_df)} rows")
    else:
        combined = pd.concat([historical_df, raw_df], ignore_index=True)
        combined = deduplicate_on_case_no(combined)
        log(f"Appended to historical data; total now: {len(combined)} rows")
    
    return combined


def compute_insights(df):
    """Compute insights and metrics from data."""
    insights = {}
    
    if df is None or len(df) == 0:
        log("WARNING: No data to compute insights")
        return insights
    
    # Overall metrics
    insights["total_cases"] = len(df)
    insights["date_range"] = f"{df['Creation Date'].min().strftime('%Y-%m-%d')} to {df['Closing Date'].max().strftime('%Y-%m-%d')}" if "Creation Date" in df.columns else "N/A"
    
    # Coaching flags
    for flag in FLAG_COLUMNS:
        if flag in df.columns:
            insights[f"{flag}_total"] = int(df[flag].sum())
            insights[f"{flag}_pct"] = round(df[flag].sum() / len(df) * 100, 1)
    
    # Cases with at least one flag
    if "Flag_TotalCoaching" in df.columns:
        flagged_cases = (df["Flag_TotalCoaching"] > 0).sum()
        insights["flagged_cases"] = int(flagged_cases)
        insights["flagged_cases_pct"] = round(flagged_cases / len(df) * 100, 1)
    
    # Time metrics
    if "Days_AtCustomer" in df.columns:
        insights["avg_days_customer"] = round(df["Days_AtCustomer"].mean(), 1)
    if "Days_AtSAP" in df.columns:
        insights["avg_days_sap"] = round(df["Days_AtSAP"].mean(), 1)
    
    # Workstream breakdown
    if "Workstream" in df.columns:
        insights["workstreams"] = df["Workstream"].value_counts().to_dict()
    
    # Priority distribution
    if "Priority" in df.columns:
        insights["priority_dist"] = df["Priority"].value_counts().to_dict()
    
    # Error categories
    if "Error Category" in df.columns:
        insights["error_categories"] = df["Error Category"].value_counts().head(10).to_dict()
    
    # Top coaching issues
    flag_counts = {flag: insights.get(f"{flag}_total", 0) for flag in FLAG_COLUMNS if flag != "Flag_TotalCoaching"}
    sorted_flags = sorted(flag_counts.items(), key=lambda x: x[1], reverse=True)
    if sorted_flags:
        insights["top_issue"] = FLAG_LABELS.get(sorted_flags[0][0], sorted_flags[0][0])
        insights["top_issue_count"] = sorted_flags[0][1]
    
    log(f"Computed insights: {insights['total_cases']} cases, {insights.get('flagged_cases', 0)} flagged")
    
    return insights


def create_insights_sheet(df, file_path):
    """Create/update the Insights sheet with current metrics."""
    if df is None or len(df) == 0:
        log("Skipping insights generation - no data")
        return
    
    insights = compute_insights(df)
    
    # Create insights dataframe
    insights_rows = []
    
    # Overall metrics
    insights_rows.append(["OVERALL METRICS", ""])
    insights_rows.append(["Total Cases", insights.get("total_cases", 0)])
    insights_rows.append(["Date Range", insights.get("date_range", "N/A")])
    insights_rows.append(["Flagged Cases", f"{insights.get('flagged_cases', 0)} ({insights.get('flagged_cases_pct', 0)}%)"])
    insights_rows.append(["", ""])
    
    # Coaching flags
    insights_rows.append(["COACHING FLAGS", "Count", "Percentage"])
    for flag in FLAG_COLUMNS:
        if flag in FLAG_LABELS:
            count = insights.get(f"{flag}_total", 0)
            pct = insights.get(f"{flag}_pct", 0)
            insights_rows.append([FLAG_LABELS[flag], count, f"{pct}%"])
    insights_rows.append(["", ""])
    
    # Time metrics
    insights_rows.append(["TIME METRICS", "Days"])
    insights_rows.append(["Avg Days @ Customer", insights.get("avg_days_customer", 0)])
    insights_rows.append(["Avg Days @ SAP", insights.get("avg_days_sap", 0)])
    insights_rows.append(["", ""])
    
    # Top issues
    insights_rows.append(["TOP COACHING ISSUE", ""])
    insights_rows.append([insights.get("top_issue", "N/A"), insights.get("top_issue_count", 0)])
    
    # Write to Excel
    wb = load_workbook(file_path)
    ws = wb[SHEET_INSIGHTS]
    ws.delete_rows(1, ws.max_row)
    
    for row_idx, row_data in enumerate(insights_rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style header rows
            if isinstance(value, str) and value.isupper() and len(value) > 3:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Update metadata
    ws_meta = wb[SHEET_METADATA]
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta["B2"] = insights.get("total_cases", 0)
    ws_meta["B3"] = "Updated"
    
    wb.save(file_path)
    log(f"Updated Insights sheet")


def backup_file():
    """Create a backup of the consolidated file."""
    if not os.path.exists(CONSOLIDATED_FILE):
        return
    
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    backup_name = f"consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    backup_path = os.path.join(BACKUP_FOLDER, backup_name)
    
    import shutil
    shutil.copy(CONSOLIDATED_FILE, backup_path)
    log(f"Created backup: {backup_path}")


def main():
    """Main processing pipeline."""
    log("=" * 60)
    log("OSS Dashboard Data Processing Started")
    log("=" * 60)
    
    try:
        # Step 1: Ensure file and sheets exist
        ensure_sheets_exist(CONSOLIDATED_FILE)
        
        # Step 2: Read raw data
        raw_df = read_raw_data()
        if raw_df is None:
            log("No raw data to process. Exiting.")
            return
        
        # Step 3: Clean raw data
        raw_df = clean_data(raw_df)
        
        # Step 4: Read historical data
        historical_df = read_historical_data()
        
        # Step 5: Append to historical
        combined_df = append_to_historical(raw_df, historical_df)
        
        # Step 6: Backup before write
        backup_file()
        
        # Step 7: Write updated historical data back
        wb = load_workbook(CONSOLIDATED_FILE)
        ws = wb[SHEET_HISTORICAL]
        ws.delete_rows(1, ws.max_row)
        
        for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(CONSOLIDATED_FILE)
        log(f"Wrote {len(combined_df)} rows to Historical Data sheet")
        
        # Step 8: Generate insights
        create_insights_sheet(combined_df, CONSOLIDATED_FILE)
        
        # Step 9: Clear raw data sheet (optional - keep for next run)
        # ws_raw = wb[SHEET_RAW_DATA]
        # ws_raw.delete_rows(1, ws_raw.max_row)
        # wb.save(CONSOLIDATED_FILE)
        # log("Cleared Raw Data sheet")
        
        log("=" * 60)
        log("Data processing completed successfully!")
        log("=" * 60)
        
    except Exception as e:
        log(f"ERROR during processing: {e}")
        raise


if __name__ == "__main__":
    main()
