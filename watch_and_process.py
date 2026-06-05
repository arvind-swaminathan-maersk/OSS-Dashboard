#!/usr/bin/env python3
"""
watch_and_process.py
====================
File system watcher that automatically processes Excel files when placed in Data folder.

This script:
1. Monitors the Data folder for new Excel files
2. When a new file is detected, automatically:
   - Reads the data
   - Appends to consolidated.xlsx Historical Data sheet
   - Deduplicates
   - Regenerates Insights sheet
   - Builds the dashboard

Usage (Local):
    python watch_and_process.py
    (Runs continuously, watching for new files)

Usage (Docker/Scheduled):
    Can be run in background or as a scheduled task

Requirements:
    pip install pandas openpyxl watchdog
"""

import pandas as pd
import os
import sys
import glob
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import shutil
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler


# Configuration
DATA_FOLDER = "Data"
CONSOLIDATED_FILE = os.path.join(DATA_FOLDER, "consolidated.xlsx")
BACKUP_FOLDER = os.path.join(DATA_FOLDER, "backups")
PROCESSED_FOLDER = os.path.join(DATA_FOLDER, "processed")
LOG_FILE = os.path.join(DATA_FOLDER, "process.log")
WATCH_EXTENSIONS = {'.xlsx', '.xls'}

# Sheet names
SHEET_RAW_DATA = "Raw Data"
SHEET_HISTORICAL = "Historical Data"
SHEET_INSIGHTS = "Insights"
SHEET_METADATA = "Metadata"

# Flag columns
FLAG_COLUMNS = [
    "Flag_HowToOrConsulting",
    "Flag_PriorityInflation",
    "Flag_CustomerDelay",
    "Flag_ComponentMismatch",
    "Flag_Reopened",
    "Flag_AutoConfirmed",
    "Flag_TotalCoaching",
]

FLAG_LABELS = {
    "Flag_HowToOrConsulting": "How-To / Consulting",
    "Flag_PriorityInflation": "Priority Inflation",
    "Flag_CustomerDelay": "Customer Delay",
    "Flag_ComponentMismatch": "Component Mismatch",
    "Flag_Reopened": "Reopened",
    "Flag_AutoConfirmed": "Auto-Confirmed",
    "Flag_TotalCoaching": "Total Coaching",
}


def log(msg):
    """Log message to console and log file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    print(log_msg)
    os.makedirs(DATA_FOLDER, exist_ok=True)
    with open(LOG_FILE, "a") as f:
        f.write(log_msg + "\n")


def ensure_folders():
    """Create necessary folders."""
    os.makedirs(DATA_FOLDER, exist_ok=True)
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    os.makedirs(PROCESSED_FOLDER, exist_ok=True)


def ensure_consolidated_exists():
    """Ensure consolidated.xlsx exists."""
    if not os.path.exists(CONSOLIDATED_FILE):
        log(f"Creating {CONSOLIDATED_FILE}")
        wb = load_workbook()
        wb.remove(wb.active)
        
        for sheet_name in [SHEET_RAW_DATA, SHEET_HISTORICAL, SHEET_INSIGHTS, SHEET_METADATA]:
            ws = wb.create_sheet(sheet_name)
            
            # Add headers to Historical Data
            if sheet_name == SHEET_HISTORICAL:
                headers = [
                    "Case No.", "Creation Date", "Closing Date", "Workstream", 
                    "Priority", "Error Category", "Days_AtCustomer", "Days_AtSAP",
                    "Flag_HowToOrConsulting", "Flag_PriorityInflation", "Flag_CustomerDelay",
                    "Flag_ComponentMismatch", "Flag_Reopened", "Flag_AutoConfirmed",
                    "Flag_TotalCoaching", "_processed_date", "_source_file"
                ]
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
        
        wb.save(CONSOLIDATED_FILE)
        log("Created consolidated.xlsx with empty sheets")


def clean_data(df):
    """Clean and standardize data types."""
    if df is None or len(df) == 0:
        return df
    
    # Clean flag columns
    for col in FLAG_COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
    
    # Clean numeric columns
    for col in ["Days_AtCustomer", "Days_AtSAP"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # Clean string columns
    for col in ["Workstream", "Priority", "Error Category", "Case No."]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
    
    # Clean date columns
    for col in ["Creation Date", "Closing Date"]:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")
    
    return df


def deduplicate_on_case_no(df):
    """Remove duplicates keeping the latest entry."""
    if df is None or len(df) == 0:
        return df
    
    if "Case No." not in df.columns:
        log("WARNING: 'Case No.' column not found - skipping deduplication")
        return df
    
    before = len(df)
    
    # Sort by Closing Date desc so latest is kept
    if "Closing Date" in df.columns:
        df["_sort_date"] = pd.to_datetime(df["Closing Date"], errors="coerce")
        df = df.sort_values("_sort_date", ascending=False, na_position="last")
        df = df.drop(columns=["_sort_date"])
    
    df = df.drop_duplicates(subset=["Case No."], keep="first")
    after = len(df)
    removed = before - after
    
    if removed > 0:
        log(f"Deduplicated: removed {removed} duplicate case(s)")
    
    return df.reset_index(drop=True)


def compute_insights(df):
    """Compute insights and metrics from data."""
    insights = {}
    
    if df is None or len(df) == 0:
        log("WARNING: No data to compute insights")
        return insights
    
    # Overall metrics
    insights["total_cases"] = len(df)
    if "Creation Date" in df.columns and "Closing Date" in df.columns:
        insights["date_range"] = f"{df['Creation Date'].min().strftime('%Y-%m-%d')} to {df['Closing Date'].max().strftime('%Y-%m-%d')}"
    else:
        insights["date_range"] = "N/A"
    
    # Coaching flags
    for flag in FLAG_COLUMNS:
        if flag in df.columns:
            insights[f"{flag}_total"] = int(df[flag].sum())
            insights[f"{flag}_pct"] = round(df[flag].sum() / len(df) * 100, 1)
    
    # Cases with flags
    if "Flag_TotalCoaching" in df.columns:
        flagged_cases = (df["Flag_TotalCoaching"] > 0).sum()
        insights["flagged_cases"] = int(flagged_cases)
        insights["flagged_cases_pct"] = round(flagged_cases / len(df) * 100, 1)
    
    # Time metrics
    if "Days_AtCustomer" in df.columns:
        insights["avg_days_customer"] = round(df["Days_AtCustomer"].mean(), 1)
    if "Days_AtSAP" in df.columns:
        insights["avg_days_sap"] = round(df["Days_AtSAP"].mean(), 1)
    
    # Top coaching issues
    flag_counts = {flag: insights.get(f"{flag}_total", 0) for flag in FLAG_COLUMNS if flag != "Flag_TotalCoaching"}
    sorted_flags = sorted(flag_counts.items(), key=lambda x: x[1], reverse=True)
    if sorted_flags:
        insights["top_issue"] = FLAG_LABELS.get(sorted_flags[0][0], sorted_flags[0][0])
        insights["top_issue_count"] = sorted_flags[0][1]
    
    return insights


def create_insights_sheet(df, file_path):
    """Create/update the Insights sheet with current metrics."""
    if df is None or len(df) == 0:
        log("Skipping insights generation - no data")
        return
    
    insights = compute_insights(df)
    
    # Create insights dataframe
    insights_rows = []
    
    # Overall metrics
    insights_rows.append(["OVERALL METRICS", ""])
    insights_rows.append(["Total Cases", insights.get("total_cases", 0)])
    insights_rows.append(["Date Range", insights.get("date_range", "N/A")])
    insights_rows.append(["Flagged Cases", f"{insights.get('flagged_cases', 0)} ({insights.get('flagged_cases_pct', 0)}%)"])
    insights_rows.append(["", ""])
    
    # Coaching flags
    insights_rows.append(["COACHING FLAGS", "Count", "Percentage"])
    for flag in FLAG_COLUMNS:
        if flag in FLAG_LABELS:
            count = insights.get(f"{flag}_total", 0)
            pct = insights.get(f"{flag}_pct", 0)
            insights_rows.append([FLAG_LABELS[flag], count, f"{pct}%"])
    insights_rows.append(["", ""])
    
    # Time metrics
    insights_rows.append(["TIME METRICS", "Days"])
    insights_rows.append(["Avg Days @ Customer", insights.get("avg_days_customer", 0)])
    insights_rows.append(["Avg Days @ SAP", insights.get("avg_days_sap", 0)])
    insights_rows.append(["", ""])
    
    # Top issues
    insights_rows.append(["TOP COACHING ISSUE", ""])
    insights_rows.append([insights.get("top_issue", "N/A"), insights.get("top_issue_count", 0)])
    
    # Write to Excel
    wb = load_workbook(file_path)
    ws = wb[SHEET_INSIGHTS]
    ws.delete_rows(1, ws.max_row)
    
    for row_idx, row_data in enumerate(insights_rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Style header rows
            if isinstance(value, str) and value.isupper() and len(value) > 3:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Update metadata
    ws_meta = wb[SHEET_METADATA]
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta["B2"] = insights.get("total_cases", 0)
    ws_meta["B3"] = "Updated"
    
    wb.save(file_path)
    log(f"Updated Insights sheet with {insights.get('total_cases', 0)} cases")


def backup_file():
    """Create a backup of the consolidated file."""
    if not os.path.exists(CONSOLIDATED_FILE):
        return
    
    backup_name = f"consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    backup_path = os.path.join(BACKUP_FOLDER, backup_name)
    
    shutil.copy(CONSOLIDATED_FILE, backup_path)
    log(f"Created backup: {backup_name}")


def process_excel_file(file_path):
    """Process a newly detected Excel file."""
    log(f"\n{'='*70}")
    log(f"Processing new file: {os.path.basename(file_path)}")
    log(f"{'='*70}")
    
    try:
        # Wait a moment for file to finish writing
        time.sleep(2)
        
        # Check file exists and is not locked
        if not os.path.exists(file_path):
            log(f"ERROR: File no longer exists: {file_path}")
            return
        
        # Step 1: Read the new data
        log(f"[1/6] Reading data from {os.path.basename(file_path)}")
        try:
            df_new = pd.read_excel(file_path, engine="openpyxl")
            rows_new = len(df_new)
            log(f"Read {rows_new} rows from file")
        except Exception as e:
            log(f"ERROR reading file: {e}")
            return
        
        # Step 2: Clean the new data
        log(f"[2/6] Cleaning data")
        df_new = clean_data(df_new)
        df_new["_processed_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df_new["_source_file"] = os.path.basename(file_path)
        
        # Step 3: Read historical data
        log(f"[3/6] Reading historical data")
        ensure_consolidated_exists()
        try:
            df_historical = pd.read_excel(CONSOLIDATED_FILE, sheet_name=SHEET_HISTORICAL)
            if len(df_historical) == 0:
                df_historical = pd.DataFrame()
        except:
            df_historical = pd.DataFrame()
        
        # Step 4: Combine and deduplicate
        log(f"[4/6] Combining and deduplicating")
        if len(df_historical) > 0:
            df_combined = pd.concat([df_historical, df_new], ignore_index=True)
        else:
            df_combined = df_new
        
        df_combined = deduplicate_on_case_no(df_combined)
        log(f"Historical data now contains {len(df_combined)} unique cases")
        
        # Step 5: Backup
        log(f"[5/6] Creating backup")
        backup_file()
        
        # Step 6: Write back to consolidated.xlsx
        log(f"[6/6] Updating consolidated.xlsx")
        wb = load_workbook(CONSOLIDATED_FILE)
        ws = wb[SHEET_HISTORICAL]
        ws.delete_rows(1, ws.max_row)
        
        for r_idx, row in enumerate(dataframe_to_rows(df_combined, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(CONSOLIDATED_FILE)
        log(f"Wrote {len(df_combined)} rows to Historical Data sheet")
        
        # Step 7: Generate insights
        create_insights_sheet(df_combined, CONSOLIDATED_FILE)
        
        # Step 8: Build dashboard
        log(f"[7/7] Building dashboard")
        try:
            import subprocess
            result = subprocess.run([sys.executable, "build_dashboard.py"], 
                                  capture_output=True, text=True, timeout=60)
            if result.returncode == 0:
                log("✅ Dashboard built successfully")
            else:
                log(f"Dashboard build had issues: {result.stderr}")
        except Exception as e:
            log(f"WARNING: Could not auto-build dashboard: {e}")
        
        # Step 9: Move processed file
        processed_name = f"{Path(file_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        processed_path = os.path.join(PROCESSED_FOLDER, processed_name)
        shutil.move(file_path, processed_path)
        log(f"Moved processed file to: {processed_name}")
        
        log(f"{'='*70}")
        log(f"✅ Processing complete!")
        log(f"{'='*70}\n")
        
    except Exception as e:
        log(f"ERROR during processing: {e}")
        import traceback
        log(traceback.format_exc())


class DataFolderHandler(FileSystemEventHandler):
    """Watch for Excel files in Data folder."""
    
    def on_created(self, event):
        """Triggered when a file is created."""
        if event.is_directory:
            return
        
        file_path = event.src_path
        file_ext = os.path.splitext(file_path)[1].lower()
        file_name = os.path.basename(file_path)
        
        # Skip consolidated.xlsx and files in subfolders
        if file_name == "consolidated.xlsx":
            return
        if "processed" in file_path or "backup" in file_path:
            return
        if file_ext not in WATCH_EXTENSIONS:
            return
        
        log(f"🔔 New file detected: {file_name}")
        process_excel_file(file_path)


def main():
    """Main watcher process."""
    log("\n" + "="*70)
    log("OSS Dashboard - File Watcher Service Started")
    log("="*70)
    log("Monitoring: Data folder for new Excel files")
    log("Action: Auto-process & append to consolidated.xlsx")
    log("Press Ctrl+C to stop")
    log("="*70 + "\n")
    
    ensure_folders()
    ensure_consolidated_exists()
    
    # Set up file system observer
    observer = Observer()
    event_handler = DataFolderHandler()
    observer.schedule(event_handler, path=DATA_FOLDER, recursive=False)
    
    try:
        observer.start()
        log(f"✅ Watcher ready. Watching {DATA_FOLDER}/ for new Excel files...\n")
        
        # Keep running
        while True:
            time.sleep(1)
    
    except KeyboardInterrupt:
        log("\n⏹ Shutting down...")
        observer.stop()
    
    observer.join()
    log("Watcher stopped.")


if __name__ == "__main__":
    main()
